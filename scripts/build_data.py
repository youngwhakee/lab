from pathlib import Path
from datetime import datetime, date
import json, re
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / 'data' / 'kee_lab_content.xlsx'
OUT = ROOT / 'assets' / 'data'
OUT.mkdir(parents=True, exist_ok=True)

def yes(v):
    if isinstance(v, bool): return v
    return str(v or '').strip().lower() in {'true','1','yes','y'}

def text(v):
    if v is None: return ''
    return str(v).strip()

def iso(v):
    if isinstance(v, (datetime, date)): return v.strftime('%Y-%m-%d')
    s=text(v)
    if not s: return ''
    for fmt in ('%Y-%m-%d','%Y/%m/%d','%Y.%m.%d'):
        try: return datetime.strptime(s,fmt).strftime('%Y-%m-%d')
        except ValueError: pass
    return s

def num(v, default=0):
    try: return int(float(v))
    except (TypeError,ValueError): return default

def normalize_role(group, label=''):
    raw = f"{text(group)} {text(label)}".strip().lower()
    compact = re.sub(r'[^a-z0-9가-힣]+', '', raw)

    if any(k in compact for k in ('phd','doctoral','doctorate','박사')):
        return 'phd'
    if any(k in compact for k in ('master','masters','ma','석사')):
        return 'ma'
    if any(k in compact for k in ('researcher','research','postdoc','연구원','연구교수')):
        return 'researcher'
    return 'other'

def rows(sheet):
    values=list(sheet.iter_rows(values_only=True))
    if not values: return []
    headers=[text(x) for x in values[0]]
    result=[]
    for row in values[1:]:
        if all(v is None or text(v)=='' for v in row): continue
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return result

def image_path(folder, filename):
    f=text(filename)
    if not f: return ''
    if f.startswith(('http://','https://','assets/')): return f
    return f'assets/uploads/{folder}/{f}'

def write(name, data):
    (OUT/f'{name}.json').write_text(json.dumps(data,ensure_ascii=False,indent=2),encoding='utf-8')

wb=load_workbook(XLSX,data_only=True)


professor={}
if 'Professor' in wb.sheetnames:
    for r in rows(wb['Professor']):
        if not yes(r.get('active')):
            continue
        professor={
            'name_ko':text(r.get('name_ko')),
            'name_en':text(r.get('name_en')),
            'title':text(r.get('title')),
            'university':text(r.get('university')),
            'department':text(r.get('department')),
            'email':text(r.get('email')),
            'office':text(r.get('office')),
            'photo':image_path('professor',r.get('photo_file')),
            'home_bio':text(r.get('home_bio')),
            'biography':text(r.get('biography')),
            'research_interests':[x.strip() for x in text(r.get('research_interests')).split('|') if x.strip()],
        }
        break
write('professor',professor)

members=[]
for r in rows(wb['Members']):
    if not yes(r.get('active')): continue
    members.append({
        'id':text(r.get('id')) or re.sub(r'[^a-z0-9]+','-',text(r.get('name_en')).lower()).strip('-'),
        'sort_order':num(r.get('sort_order'),999), 'name_ko':text(r.get('name_ko')), 'name_en':text(r.get('name_en')),
        'role_group':normalize_role(r.get('role_group'), r.get('role_label')), 'role_label':text(r.get('role_label')),
        'research_interests':text(r.get('research_interests')), 'email':text(r.get('email')),
        'profile_url':text(r.get('profile_url')), 'photo':image_path('members',r.get('photo_file')),
    })
members.sort(key=lambda x:(x['sort_order'],x['name_ko'] or x['name_en']))
write('members',members)

alumni=[]
for r in rows(wb['Alumni']):
    if not yes(r.get('active')): continue
    alumni.append({'id':text(r.get('id')),'name_ko':text(r.get('name_ko')),'name_en':text(r.get('name_en')),'degree':text(r.get('degree')),'graduation_year':num(r.get('graduation_year')),'affiliation':text(r.get('affiliation')),'position':text(r.get('position')),'note':text(r.get('note'))})
alumni.sort(key=lambda x:(-x['graduation_year'],x['name_ko'] or x['name_en']))
write('alumni',alumni)

publications=[]
for r in rows(wb['Publications']):
    if not yes(r.get('active')): continue
    publications.append({'id':text(r.get('id')),'featured':yes(r.get('featured')),'year':num(r.get('year')),'type':text(r.get('type')).lower(),'title':text(r.get('title')),'authors':text(r.get('authors')),'venue':text(r.get('venue')),'volume_issue':text(r.get('volume_issue')),'doi_url':text(r.get('doi_url')),'pdf_url':text(r.get('pdf_url')),'keywords':text(r.get('keywords'))})
publications.sort(key=lambda x:(-x['year'],not x['featured'],x['title']))
write('publications',publications)

news=[]
for r in rows(wb['News']):
    if not yes(r.get('active')): continue
    news.append({'id':text(r.get('id')),'featured':yes(r.get('featured')),'date':iso(r.get('date')),'category':text(r.get('category')).lower(),'title':text(r.get('title')),'summary':text(r.get('summary')),'body':text(r.get('body')),'image':image_path('news',r.get('image_file')),'link_url':text(r.get('link_url'))})
news.sort(key=lambda x:(x['date'],x['featured']),reverse=True)
write('news',news)

gallery=[]
for r in rows(wb['Gallery']):
    if not yes(r.get('active')): continue
    gallery.append({'id':text(r.get('id')),'featured':yes(r.get('featured')),'date':iso(r.get('date')),'category':text(r.get('category')).lower(),'title':text(r.get('title')),'caption':text(r.get('caption')),'image':image_path('gallery',r.get('image_file')),'album':text(r.get('album'))})
gallery.sort(key=lambda x:(x['date'],x['featured']),reverse=True)
write('gallery',gallery)

print(f"Built: professor={'yes' if professor else 'no'}, {len(members)} members, {len(alumni)} alumni, {len(publications)} publications, {len(news)} news, {len(gallery)} photos")
